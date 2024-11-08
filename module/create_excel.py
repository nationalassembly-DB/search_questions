"""
입력받은 북마크 리스트를 통해 엑셀 파일을 만듭니다. 북마크의 단계를 기준으로 추출합니다
"""

import os
from natsort import natsorted

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from module.data import committee_dict, organization_dict, person_dict
from module.create_log import logging
from module.extract_bookmark import extract_bookmark


def write_excel(wb, input_path, output_path, book_level):  # pylint: disable=R0914
    """정리된 북마크 리스트를 통해 엑셀로 변환합니다. """
    ws = wb.active

    for root, _, files in os.walk(input_path):
        for file in natsorted(files):
            if not file.lower().endswith('.pdf'):
                continue
            cmt = _extract_cmt(file)
            org = _extract_org(file)

            file_path = os.path.join(root, file)
            last_row = ws.max_row
            tmp = 1
            for item in extract_bookmark(file_path):
                if len(item) > 1 and item['level'] == book_level:
                    ws.cell(row=last_row + tmp, column=2, value=org)
                    ws.cell(row=last_row + tmp, column=3,
                            value=organization_dict[org] if org in organization_dict else None)
                    ws.cell(row=last_row + tmp, column=4, value=cmt)
                    ws.cell(row=last_row + tmp, column=5,
                            value=committee_dict[cmt] if cmt in committee_dict else None)
                    ws.cell(row=last_row + tmp, column=11, value=file)
                    if book_level > 1:
                        ws.cell(row=last_row + tmp, column=6,
                                value=item['parent']['title'])
                        ws.cell(row=last_row + tmp, column=7,
                                value=person_dict[item['parent']['title']]
                                if item['parent']['title'] in person_dict else None)
                    ws.cell(row=last_row + tmp, column=9,
                            value=item['title'])
                    tmp += 1

    wb.save(output_path)


def has_header(wb, output_path):
    """엑셀 header가 존재하는지 확인합니다. 존재하지 않을 경우 새로 생성합니다"""
    ws = wb.active
    first_row = ws[1]
    header_exists = any(cell.value for cell in first_row)

    if not header_exists:
        headers = ['일련번호', '기관명', '기관코드', '위원회명', '위원회 코드',
                   '위원(의원)명', '위원(의원) 코드', '질의유형', '질의', '답변 책자 파일명', '파일명']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        fill_color = PatternFill(start_color='4f81bd',
                                 end_color='4f81bd', fill_type='solid')
        for col in range(1, 12):
            ws.cell(row=1, column=col).fill = fill_color
    wb.save(output_path)
    return wb


def load_excel(output_path):
    """엑셀을 불러옵니다. 파일이 없는 경우 새로 생성됩니다"""
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        try:
            output_excel = output_path
            wb = Workbook()
            wb.save(output_excel)
            wb = load_workbook(output_excel)
        except Exception:  # pylint: disable=W0703
            e = "엑셀 파일 생성 오류"
            logging(e, '', output_path)

    return has_header(wb, output_path)


def _extract_cmt(filename):
    # 파일명에서 위원회 이름 추출
    first_underscore_index = filename.find('_')
    second_underscore_index = filename.find(
        '_', first_underscore_index + 1)
    if first_underscore_index != -1 and second_underscore_index != -1:
        cmt = filename[first_underscore_index +
                       1:second_underscore_index]
    else:
        cmt = ""

    return cmt


def _extract_org(filename):
    # 파일명에서 가장 바깥 괄호를 기준으로 기관 이름 추출 (뒤에서부터 탐색)
    stack = []
    end = None
    org = ""

    for i in range(len(filename) - 1, -1, -1):
        char = filename[i]

        if char == ')':
            stack.append(i)
            if len(stack) == 1:
                end = i
        elif char == '(':
            stack.pop()
            if len(stack) == 0:
                org = filename[i + 1:end]
                break

    return org if org else ""
