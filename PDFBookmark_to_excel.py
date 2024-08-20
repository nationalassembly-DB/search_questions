import os
import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def extract_bookmark(pdf_path):
    doc = fitz.open(pdf_path)
    toc = doc.get_toc(simple=False)
    return toc


def write_excel(wb, input_path, output_path):
    try:
        ws = wb.active
        for root, _, files in os.walk(input_path):
            for file in files:
                file_path = os.path.join(root, file)
                bookmark_list = extract_bookmark(file_path)
                last_row = ws.max_row
                for idx, item in enumerate(bookmark_list):
                    _, title, _, _ = item
                    ws.cell(row=last_row + idx + 1, column=9, value=title)
        wb.save(output_path)
    except Exception as e:
        logging(e, input_path, output_path)


def has_header(wb, output_path):
    ws = wb.active
    first_row = ws[1]
    header_exists = any(cell.value for cell in first_row)

    if not header_exists:
        headers = ['일련번호', '기관명', '기관코드', '위원회명', '위원회 코드',
                   '위원(의원)명', '위원(의원) 코드', '질의유형', '질의', '답변 책자 파일명']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        fill_color = PatternFill(start_color='4f81bd',
                                 end_color='4f81bd', fill_type='solid')
        for col in range(1, 11):
            ws.cell(row=1, column=col).fill = fill_color
    wb.save(output_path)
    return wb


def load_excel(output_path):
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
    else:
        try:
            output_excel = output_path
            wb = Workbook()
            wb.save(output_excel)
            wb = load_workbook(output_excel)
        except Exception:
            e = "엑셀 파일 생성 오류"
            logging(e, '', output_path)

    return has_header(wb, output_path)


def logging(e, input_path, output_path):
    output_path = os.path.dirname(output_path)
    log_file_path = os.path.join(output_path, 'log.txt')
    with open(log_file_path, 'a', encoding='utf-8') as log_file:
        log_file.write(f'{e} {input_path} {output_path}\n')


def main():
    input_path = input("PDF파일이 존재하는 폴더 경로를 입력하세요 : ").strip()
    output_path = input("저장할 엑셀 파일 경로를 입력하세요 : ").strip()

    if not os.path.isdir(input_path):
        print("입력 폴더의 경로를 다시 한번 확인하세요")
        return main
    if not str(output_path).endswith('.xlsx'):
        output_path = output_path + '.xlsx'
    write_excel(load_excel(output_path), input_path, output_path)


if __name__ == "__main__":
    main()
